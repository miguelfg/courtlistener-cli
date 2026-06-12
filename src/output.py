"""Output formatting utilities for JSON, CSV, XLSX"""

import json
import csv
import click
from pathlib import Path
from typing import Any, Dict, List, Optional
from datetime import datetime


def print_to_screen(data: Any):
    """Print results to screen if --screen flag is set"""
    ctx = click.get_current_context(silent=True)
    if ctx and ctx.obj and ctx.obj.get('screen'):
        # If it's a list of results, we might want to wrap it in a dict for consistency
        if isinstance(data, list):
            output = {"results": data, "count": len(data)}
        else:
            output = data
        click.echo(json.dumps(output, indent=2))


def flatten_dict(
    d: Dict[str, Any],
    parent_key: str = '',
    sep: str = '.',
    concat_list_fields_char: Optional[str] = ';',
) -> Dict[str, Any]:
    """Flatten nested dictionary using dot notation.

    Args:
        concat_list_fields_char: Character used to join list values into a
            string. When ``None`` or empty, lists are serialised as JSON.
    """
    items = []
    for k, v in d.items():
        new_key = f"{parent_key}{sep}{k}" if parent_key else k
        if isinstance(v, dict):
            items.extend(flatten_dict(v, new_key, sep=sep, concat_list_fields_char=concat_list_fields_char).items())
        elif isinstance(v, list):
            if concat_list_fields_char:
                items.append((new_key, concat_list_fields_char.join(str(i) for i in v)))
            else:
                items.append((new_key, json.dumps(v)))
        else:
            items.append((new_key, v))
    return dict(items)


def save_json(data: Any, output_path: Path, include_timestamp: bool = False, filename_stem: str = "results") -> Path:
    """Save data as JSON"""
    print_to_screen(data)
    if include_timestamp:
        filename = f"{filename_stem}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
    else:
        filename = f"{filename_stem}.json"
    
    filepath = output_path / filename
    with open(filepath, 'w') as f:
        json.dump(data, f, indent=2)
    
    return filepath


def save_csv(
    data: List[Dict],
    output_path: Path,
    include_timestamp: bool = False,
    filename_stem: str = "results",
    concat_list_fields_char: Optional[str] = ';',
) -> Path:
    """Save data as CSV"""
    print_to_screen(data)
    if include_timestamp:
        filename = f"{filename_stem}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    else:
        filename = f"{filename_stem}.csv"

    filepath = output_path / filename

    # Flatten nested dictionaries
    flattened = [flatten_dict(row, concat_list_fields_char=concat_list_fields_char) for row in data]
    
    if flattened:
        with open(filepath, 'w', newline='') as f:
            writer = csv.DictWriter(f, fieldnames=flattened[0].keys())
            writer.writeheader()
            writer.writerows(flattened)
    
    return filepath


def save_xlsx(
    data: List[Dict],
    output_path: Path,
    include_timestamp: bool = False,
    filename_stem: str = "results",
    concat_list_fields_char: Optional[str] = ';',
) -> Path:
    """Save data as XLSX"""
    print_to_screen(data)
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise ImportError("openpyxl required for XLSX export. Install: pip install openpyxl")
    
    if include_timestamp:
        filename = f"{filename_stem}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    else:
        filename = f"{filename_stem}.xlsx"
    
    filepath = output_path / filename
    
    # Flatten nested dictionaries
    flattened = [flatten_dict(row, concat_list_fields_char=concat_list_fields_char) for row in data]

    if not flattened:
        return filepath
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Results"
    
    # Write header
    headers = list(flattened[0].keys())
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    # Write data
    for row_idx, row_data in enumerate(flattened, 2):
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=row_idx, column=col_idx, value=row_data.get(header, ''))
    
    # Auto-size columns
    for col in ws.columns:
        max_length = 0
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)
    
    wb.save(filepath)
    return filepath
