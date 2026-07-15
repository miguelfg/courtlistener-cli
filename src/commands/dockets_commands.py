"""Click commands for dockets resource"""

import csv
import io
import re
import click
import json
import httpx
import logging
from ..client import CourtListenerClient
from ..output import save_json, save_csv, save_xlsx
from ..pagination import paginate_endpoint
from pathlib import Path
from typing import List, Optional

logger = logging.getLogger(__name__)

_STORAGE_BASE = "https://storage.courtlistener.com/"


def _safe_folder_name(name: str) -> str:
    """Strip characters that are illegal in directory names on common OSes."""
    return re.sub(r'[<>:"/\\|?*\x00-\x1f]', "_", name).strip()


def _download_docs_folder_name(
    docket_id: int, docket: dict, folder_name_mode: str
) -> Optional[str]:
    """Build the subfolder name for download-docs, or None to use the output dir directly."""
    case_name = (
        docket.get("case_name")
        or docket.get("case_name_short")
        or f"docket_{docket_id}"
    )
    docket_number = docket.get("docket_number") or str(docket_id)

    if folder_name_mode == "none":
        return None
    if folder_name_mode == "docket-id":
        return str(docket_id)
    if folder_name_mode == "docket-number":
        return docket_number
    if folder_name_mode == "case-name":
        return case_name
    return f"{case_name} ; {docket_number}"


def _read_batch_values(input_file: Path, column: str) -> List[str]:
    """Read query values from a CSV/XLSX file column."""
    suffix = input_file.suffix.lower()
    if suffix == ".csv":
        with open(input_file, "r", newline="") as f:
            reader = csv.DictReader(f)
            if not reader.fieldnames or column not in reader.fieldnames:
                raise ValueError(f"Column '{column}' not found in CSV file")
            return [
                str(row[column]).strip()
                for row in reader
                if row.get(column) is not None and str(row[column]).strip()
            ]

    if suffix == ".xlsx":
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise ImportError(
                "openpyxl required for XLSX input. Install: pip install openpyxl"
            )

        wb = load_workbook(input_file, read_only=True, data_only=True)
        ws = wb.worksheets[0]
        rows = ws.iter_rows(values_only=True)
        headers = next(rows, None)
        if not headers:
            raise ValueError("Input XLSX file has no header row")
        header_index = {
            str(h).strip(): idx for idx, h in enumerate(headers) if h is not None
        }
        if column not in header_index:
            raise ValueError(f"Column '{column}' not found in XLSX first sheet")

        idx = header_index[column]
        values = []
        for row in rows:
            if row is None or idx >= len(row):
                continue
            value = row[idx]
            if value is None:
                continue
            value = str(value).strip()
            if value:
                values.append(value)
        return values

    raise ValueError("Unsupported input format. Use CSV or XLSX.")


@click.group()
def dockets():
    """Manage case dockets"""
    pass


@dockets.command("list")
@click.argument(
    "input_file", required=False, type=click.Path(exists=True, dir_okay=False)
)
@click.option(
    "--column",
    default=None,
    help="Column name in input CSV/XLSX containing docket numbers or IDs",
)
@click.option(
    "--limit",
    default=20,
    type=int,
    help="Total results to export per request; use 0 with --max-pages 0 to export all results",
)
@click.option(
    "--max-pages", default=10, type=int, help="Maximum pages to fetch (0 = no page cap)"
)
@click.option("--offset", default=0, help="Pagination offset")
@click.option("--court", help="Filter by court ID")
@click.option("--docket-number", help="Filter by docket number")
@click.option("--case-name", help="Filter by case name")
@click.option(
    "--format",
    "output_format",
    default="json",
    type=click.Choice(["json", "csv", "xlsx"]),
)
@click.option("--output", "output_path", default="./output", type=click.Path())
def list_dockets(
    input_file,
    column,
    limit,
    max_pages,
    offset,
    court,
    docket_number,
    case_name,
    output_format,
    output_path,
):
    """List case dockets or batch query from input CSV/XLSX."""
    client = CourtListenerClient()

    try:
        output_dir = Path(output_path)
        output_dir.mkdir(exist_ok=True)

        if input_file:
            if not column:
                raise ValueError("--column is required when input_file is provided")

            query_values = _read_batch_values(Path(input_file), column)
            results = []
            errors = 0
            pages_fetched = 0

            for idx, query_value in enumerate(query_values, start=1):
                click.echo(f"→ Query {idx}/{len(query_values)}: {query_value}")
                try:
                    if column.lower() in {"id", "docket_id", "docket"}:
                        docket = client.get(f"/dockets/{query_value}/")
                        pages_fetched += 1
                        if isinstance(docket, dict):
                            docket["_query_value"] = query_value
                            results.append(docket)
                            click.echo("  → Item page 1: +1 dockets (accumulated 1/1)")
                    else:
                        params = {
                            "page_size": 100 if limit == 0 else max(limit, 1),
                            "docket_number": query_value,
                        }
                        if court:
                            params["court"] = court
                        if docket_number:
                            params["docket_number"] = docket_number
                        if case_name:
                            params["case_name"] = case_name
                        docket_page_data = paginate_endpoint(
                            fetch_page=lambda request_params: client.get(
                                "/dockets/", params=request_params
                            ),
                            initial_params=params,
                            limit=limit,
                            max_pages=max_pages,
                            progress_logger=lambda page, page_count, accumulated, target: (
                                click.echo(
                                    f"  → Item page {page}: +{page_count} dockets "
                                    f"(accumulated {accumulated}/{target if target is not None else 'all'})"
                                )
                            ),
                        )
                        pages_fetched += docket_page_data.get("pages_fetched", 0)
                        for docket in docket_page_data.get("results", []):
                            if isinstance(docket, dict):
                                docket["_query_value"] = query_value
                            results.append(docket)
                except Exception:
                    errors += 1

            if output_format == "json":
                filepath = save_json(
                    {
                        "query_count": len(query_values),
                        "result_count": len(results),
                        "pages_fetched": pages_fetched,
                        "error_count": errors,
                        "results": results,
                    },
                    output_dir,
                )
            elif output_format == "csv":
                filepath = save_csv(results, output_dir)
            else:  # xlsx
                filepath = save_xlsx(results, output_dir)

            click.echo(
                f"✓ Processed {len(query_values)} query values from {input_file}"
            )
            click.echo(f"✓ Retrieved {len(results)} dockets")
            click.echo(f"✓ Fetched {pages_fetched} page(s)")
            if errors:
                click.echo(f"✗ Errors: {errors}")
            click.echo(f"✓ Saved to {filepath}")
        else:
            params = {"page_size": 100 if limit == 0 else max(limit, 1)}
            if court:
                params["court"] = court
            if docket_number:
                params["docket_number"] = docket_number
            if case_name:
                params["case_name"] = case_name

            output_data = paginate_endpoint(
                fetch_page=lambda request_params: client.get(
                    "/dockets/", params=request_params
                ),
                initial_params=params,
                limit=limit,
                max_pages=max_pages,
                progress_logger=lambda page, page_count, accumulated, target: (
                    click.echo(
                        f"→ Page {page}: +{page_count} dockets "
                        f"(accumulated {accumulated}/{target if target is not None else 'all'})"
                    )
                ),
            )

            if "results" in output_data:
                if output_format == "json":
                    filepath = save_json(output_data, output_dir)
                elif output_format == "csv":
                    filepath = save_csv(output_data["results"], output_dir)
                else:  # xlsx
                    filepath = save_xlsx(output_data["results"], output_dir)

                click.echo(f"✓ Found {output_data.get('count', 0)} total dockets")
                click.echo(f"✓ Exported {output_data.get('returned_count', 0)} dockets")
                click.echo(f"✓ Fetched {output_data.get('pages_fetched', 0)} page(s)")
                click.echo(f"✓ Saved to {filepath}")
            else:
                click.echo("No dockets found")
    except ValueError as e:
        click.echo(f"Error: {e}", err=True)
        raise SystemExit(1)
    except Exception as e:
        click.echo(f"Error: {e}", err=True)
        raise SystemExit(1)


@dockets.command("get")
@click.argument("docket_id", type=int)
@click.option("--format", "output_format", default="json", type=click.Choice(["json"]))
def get_docket(docket_id, output_format):
    """Get docket details by ID"""
    client = CourtListenerClient()

    try:
        result = client.get(f"/dockets/{docket_id}/")

        if output_format == "json":
            click.echo(json.dumps(result, indent=2))
        else:
            click.echo(f"Docket {docket_id} details retrieved")

    except Exception as e:
        click.echo(f"Error: {e}", err=True)
        raise SystemExit(1)


@dockets.command("entries")
@click.argument("docket_id", type=int)
@click.option(
    "--limit",
    default=20,
    type=int,
    help="Total results to export; use 0 with --max-pages 0 to export all results",
)
@click.option(
    "--max-pages", default=10, type=int, help="Maximum pages to fetch (0 = no page cap)"
)
@click.option(
    "--format",
    "output_format",
    default="json",
    type=click.Choice(["json", "csv", "xlsx"]),
)
@click.option("--output", "output_path", default="./output", type=click.Path())
@click.option(
    "--filename",
    "filename_stem",
    default="results",
    help="Output filename stem (without extension)",
)
@click.option(
    "--slim",
    is_flag=True,
    default=False,
    help="Also export a slim version with key fields only",
)
def get_docket_entries(
    docket_id, limit, max_pages, output_format, output_path, filename_stem, slim
):
    """Get entries for a specific docket"""
    client = CourtListenerClient()

    params = {"docket": docket_id, "page_size": 100 if limit == 0 else max(limit, 1)}

    try:
        output_data = paginate_endpoint(
            fetch_page=lambda request_params: client.get(
                "/docket-entries/", params=request_params
            ),
            initial_params=params,
            limit=limit,
            max_pages=max_pages,
            progress_logger=lambda page, page_count, accumulated, target: click.echo(
                f"→ Page {page}: +{page_count} entries "
                f"(accumulated {accumulated}/{target if target is not None else 'all'})"
            ),
        )

        # Export results
        output_dir = Path(output_path)
        output_dir.mkdir(exist_ok=True)

        if "results" in output_data:
            if output_format == "json":
                filepath = save_json(
                    output_data, output_dir, filename_stem=filename_stem
                )
            elif output_format == "csv":
                filepath = save_csv(
                    output_data["results"], output_dir, filename_stem=filename_stem
                )
            else:  # xlsx
                filepath = save_xlsx(
                    output_data["results"], output_dir, filename_stem=filename_stem
                )

            click.echo(f"✓ Found {output_data.get('count', 0)} total entries")
            click.echo(f"✓ Exported {output_data.get('returned_count', 0)} entries")
            click.echo(f"✓ Fetched {output_data.get('pages_fetched', 0)} page(s)")
            click.echo(f"✓ Saved to {filepath}")

            if slim:
                from ..reducers import slim_entries

                slim_data = slim_entries(output_data["results"])
                slim_stem = f"{filename_stem}_slim"
                if output_format == "json":
                    slim_path = save_json(
                        {"count": output_data.get("count"), "results": slim_data},
                        output_dir,
                        filename_stem=slim_stem,
                    )
                elif output_format == "csv":
                    slim_path = save_csv(slim_data, output_dir, filename_stem=slim_stem)
                else:
                    slim_path = save_xlsx(
                        slim_data, output_dir, filename_stem=slim_stem
                    )
                click.echo(f"✓ Slim export saved to {slim_path}")
        else:
            click.echo("No entries found")
    except ValueError as e:
        click.echo(f"Error: {e}", err=True)
        raise SystemExit(1)
    except Exception as e:
        click.echo(f"Error: {e}", err=True)
        raise SystemExit(1)


@dockets.command("count")
@click.option("--court", help="Filter by court ID")
@click.option("--case-name", help="Filter by case name")
def count_dockets(court, case_name):
    """Return total matching dockets count"""
    client = CourtListenerClient()

    params = {"page_size": 1}
    if court:
        params["court"] = court
    if case_name:
        params["case_name"] = case_name

    try:
        result = client.get("/dockets/", params=params)
        click.echo(result.get("count", 0))
    except Exception as e:
        click.echo(f"Error: {e}", err=True)
        raise SystemExit(1)


@dockets.command("download-docs")
@click.argument("docket_id", type=int)
@click.option(
    "--output",
    "output_path",
    default="./output",
    type=click.Path(),
    help="Parent directory for downloaded files",
)
@click.option(
    "--manifest",
    "manifest_format",
    default="xlsx",
    type=click.Choice(["xlsx", "csv"]),
    help="Format for the manifest spreadsheet",
)
@click.option(
    "--all-docs",
    "all_docs",
    is_flag=True,
    default=False,
    help="Also fetch unavailable docs (slower — paginates all recap documents)",
)
@click.option(
    "--folder-name-mode",
    default="case-name-number",
    type=click.Choice(
        ["case-name-number", "case-name", "docket-number", "docket-id", "none"]
    ),
    help="Choose how the download folder is named; use none to write directly into --output",
)
def download_docs(docket_id, output_path, manifest_format, all_docs, folder_name_mode):
    """Download all free PDFs for a docket and generate a manifest.

    By default only fetches available docs (is_available=True), which is fast
    and avoids rate-limit throttling. Use --all-docs to include unavailable ones.
    """
    client = CourtListenerClient()

    try:
        # --- 1. Fetch docket metadata for folder naming ---
        click.echo(f"→ Fetching docket {docket_id} metadata…")
        docket = client.get(f"/dockets/{docket_id}/")
        folder_name = _download_docs_folder_name(docket_id, docket, folder_name_mode)

        output_dir = Path(output_path)
        output_dir.mkdir(parents=True, exist_ok=True)
        case_dir = (
            output_dir
            if folder_name is None
            else output_dir / _safe_folder_name(folder_name)
        )
        case_dir.mkdir(parents=True, exist_ok=True)
        click.echo(f"→ Case folder: {case_dir}")

        # --- 2. Fetch doc list via CSV export (fast) or API pagination (fallback) ---
        csv_export_url = f"https://www.courtlistener.com/docket/{docket_id}/download/"
        doc_rows = _try_csv_export(csv_export_url, client, case_dir)

        if doc_rows is None:
            click.echo("→ CSV export unavailable, falling back to API pagination…")
            doc_rows = _doc_rows_from_api(docket_id, client, case_dir, not all_docs)
        else:
            if not all_docs:
                doc_rows = [r for r in doc_rows if r["is_available"]]
            click.echo(f"→ {len(doc_rows)} recap documents loaded from CSV export")

        docs_path = save_xlsx(doc_rows, case_dir, filename_stem=f"docs_{docket_id}")
        click.echo(f"→ Docs metadata saved to {docs_path}")

        # --- 4. Download available PDFs ---
        downloaded = 0
        skipped = 0
        download_headers = {"User-Agent": "courtlistener-cli/1.0.0"}

        for row in doc_rows:
            if not row["is_available"] or not row["pdf_filename"]:
                skipped += 1
                row["download_status"] = "not_available"
                continue

            dest = case_dir / row["pdf_filename"]
            if dest.exists():
                click.echo(f"  ✓ Already exists: {row['pdf_filename']}")
                row["download_status"] = "already_exists"
                row["local_file_exists"] = True
                continue

            try:
                with httpx.stream(
                    "GET",
                    row["download_url"],
                    headers=download_headers,
                    timeout=60,
                    follow_redirects=True,
                ) as r:
                    r.raise_for_status()
                    with open(dest, "wb") as f:
                        for chunk in r.iter_bytes(chunk_size=65536):
                            f.write(chunk)
                click.echo(f"  ↓ {row['pdf_filename']}  (doc {row['document_number']})")
                downloaded += 1
                row["download_status"] = "downloaded"
                row["local_file_exists"] = True
            except Exception as exc:
                logger.warning("Failed to download %s: %s", row["pdf_filename"], exc)
                click.echo(f"  ✗ Failed: {row['pdf_filename']} — {exc}", err=True)
                skipped += 1
                row["download_status"] = "failed"

        # --- 5. Write manifest ---
        manifest_rows = [
            {
                "pdf_filename": r["pdf_filename"],
                "case_folder": folder_name,
                "document_number": r["document_number"],
                "attachment_number": r["attachment_number"],
                "filing_description": r["filing_description"],
                "is_available": r["is_available"],
                "download_status": r["download_status"],
                "local_file_exists": r["local_file_exists"],
                "download_url": r["download_url"],
                "recap_doc_id": r["recap_doc_id"],
            }
            for r in doc_rows
        ]

        if manifest_rows:
            manifest_stem = f"manifest_{docket_id}"
            if manifest_format == "xlsx":
                manifest_path = save_xlsx(
                    manifest_rows, case_dir, filename_stem=manifest_stem
                )
            else:
                manifest_path = save_csv(
                    manifest_rows, case_dir, filename_stem=manifest_stem
                )
            click.echo(f"✓ Manifest saved to {manifest_path}")
        else:
            click.echo("No documents found — manifest not written")

        click.echo(f"✓ Downloaded: {downloaded}  |  Skipped/unavailable: {skipped}")

    except Exception as e:
        click.echo(f"Error: {e}", err=True)
        raise SystemExit(1)


def _try_csv_export(
    url: str, client: "CourtListenerClient", case_dir: Path
) -> List[dict] | None:
    """Fetch the CourtListener docket CSV export. Returns doc_rows or None if unavailable.

    The /docket/{id}/download/ endpoint requires a browser session cookie, not
    the API token. Set COURTLISTENER_SESSION_ID in .env (DevTools → Application
    → Cookies → sessionid).
    """
    from ..config import config

    session_id = config.session_id
    if not session_id:
        return None

    docket_url = url.replace("/download/", "/")
    headers = {
        "User-Agent": "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/148.0.0.0 Safari/537.36",
        "Accept": "*/*",
        "hx-request": "true",
        "hx-current-url": docket_url,
        "Referer": docket_url,
    }
    cookies = {"sessionid": session_id}

    try:
        response = httpx.get(
            url, headers=headers, cookies=cookies, timeout=30, follow_redirects=True
        )
        if response.status_code in (401, 403, 404):
            return None
        response.raise_for_status()
    except httpx.HTTPStatusError:
        return None

    # Save raw CSV into the case folder
    csv_filename = response.headers.get("content-disposition", "")
    csv_filename = (
        csv_filename.split("filename=")[-1].strip('"')
        if "filename=" in csv_filename
        else "docket_export.csv"
    )
    csv_path = case_dir / csv_filename
    csv_path.write_bytes(response.content)
    click.echo(f"→ CSV export saved to {csv_path}")

    rows = []
    for rec in csv.DictReader(io.StringIO(response.text)):
        filepath = rec.get("recapdocument_filepath_local", "").strip()
        pdf_filename = Path(filepath).name if filepath else ""
        rows.append(
            {
                "recap_doc_id": rec.get("recapdocument_id", ""),
                "entry_number": rec.get("docketentry_entry_number", ""),
                "filing_description": rec.get("docketentry_description", ""),
                "document_number": rec.get("recapdocument_document_number", ""),
                "attachment_number": rec.get("recapdocument_attachment_number", ""),
                "doc_description": rec.get("recapdocument_description", ""),
                "is_available": rec.get("recapdocument_is_available", "").strip()
                == "True",
                "is_free_on_pacer": rec.get(
                    "recapdocument_is_free_on_pacer", ""
                ).strip()
                == "True",
                "page_count": rec.get("recapdocument_page_count", ""),
                "file_size": rec.get("recapdocument_file_size", ""),
                "date_upload": rec.get("recapdocument_date_upload", ""),
                "pdf_filename": pdf_filename,
                "download_url": filepath,  # already a full URL in the CSV export
                "local_file_exists": (case_dir / pdf_filename).exists()
                if pdf_filename
                else False,
                "download_status": "",
            }
        )
    return rows


def _doc_rows_from_api(
    docket_id: int, client: "CourtListenerClient", case_dir: Path, available_only: bool
) -> List[dict]:
    """Fallback: build doc_rows by paginating the recap-documents API."""
    click.echo("→ Fetching recap documents via API…")
    recap_params: dict = {"docket_entry__docket": docket_id, "page_size": 100}
    if available_only:
        recap_params["is_available"] = True

    docs_data = paginate_endpoint(
        fetch_page=lambda params: client.get("/recap-documents/", params=params),
        initial_params=recap_params,
        limit=0,
        max_pages=0,
        progress_logger=lambda page, count, acc, target: click.echo(
            f"  → Page {page}: +{count} docs (total {acc})"
        ),
    )
    fetched_docs = docs_data.get("results", [])
    click.echo(f"→ {len(fetched_docs)} recap documents fetched")

    rows = []
    for doc in fetched_docs:
        filepath_local = doc.get("filepath_local") or ""
        pdf_filename = Path(filepath_local).name if filepath_local else ""
        rows.append(
            {
                "recap_doc_id": doc.get("id", ""),
                "entry_number": doc.get("document_number", ""),
                "filing_description": (doc.get("description") or "").strip(),
                "document_number": doc.get("document_number", ""),
                "attachment_number": doc.get("attachment_number", ""),
                "doc_description": (doc.get("description") or "").strip(),
                "is_available": bool(doc.get("is_available")),
                "is_free_on_pacer": bool(doc.get("is_free_on_pacer")),
                "page_count": doc.get("page_count", ""),
                "file_size": doc.get("file_size", ""),
                "date_upload": doc.get("date_upload", ""),
                "pdf_filename": pdf_filename,
                "download_url": (_STORAGE_BASE + filepath_local)
                if filepath_local
                else "",
                "local_file_exists": (case_dir / pdf_filename).exists()
                if pdf_filename
                else False,
                "download_status": "",
            }
        )
    return rows


@dockets.command("parties")
@click.argument("docket_id", type=int)
@click.option(
    "--format",
    "output_format",
    default="xlsx",
    type=click.Choice(["json", "csv", "xlsx"]),
)
@click.option("--output", "output_path", default="./output", type=click.Path())
@click.option(
    "--filename",
    "filename_stem",
    default=None,
    help="Output filename stem (without extension); defaults to parties_{docket_id}",
)
def get_parties(docket_id, output_format, output_path, filename_stem):
    """Get parties and criminal charges for a docket.

    Fetches all parties (defendants, attorneys, etc.) for DOCKET_ID and
    includes their criminal counts (charge description, disposition) when
    present — i.e. the charges/citations table from the Parties & Attorneys tab.
    """
    client = CourtListenerClient()

    try:
        click.echo(f"→ Fetching parties for docket {docket_id}…")
        parties_data = paginate_endpoint(
            fetch_page=lambda params: client.get("/parties/", params=params),
            initial_params={"docket": docket_id},
            limit=0,
            max_pages=0,
            progress_logger=lambda page, count, acc, target: click.echo(
                f"  → Page {page}: +{count} parties (total {acc})"
            ),
        )
        parties = parties_data.get("results", [])

        if not parties:
            click.echo("No parties found for this docket")
            return

        # Flatten criminal_counts into one row per count so charges are readable.
        # Real API structure: party.party_types[].criminal_counts[] and .criminal_complaints[]
        # Charge field is `name` (e.g. "18:371 (Conspiracy)(1ss)"), not `description`.
        rows: List[dict] = []
        for party in parties:
            party_types = party.get("party_types") or []
            attorneys = party.get("attorneys") or []

            # Collect attorney IDs (API returns URL references, not name strings)
            attorney_ids = list(
                {a.get("attorney_id") for a in attorneys if a.get("attorney_id")}
            )

            for pt in party_types:
                criminal_counts = pt.get("criminal_counts") or []
                criminal_complaints = pt.get("criminal_complaints") or []

                base = {
                    "party_name": party.get("name", ""),
                    "party_type": pt.get("name", ""),
                    "date_terminated": pt.get("date_terminated") or "",
                    "highest_offense_opening": pt.get(
                        "highest_offense_level_opening", ""
                    ),
                    "highest_offense_terminated": pt.get(
                        "highest_offense_level_terminated", ""
                    ),
                    "attorney_ids": "; ".join(str(i) for i in attorney_ids),
                    "criminal_complaints": "; ".join(
                        c.get("plain_text", "")
                        for c in criminal_complaints
                        if c.get("plain_text")
                    ),
                }

                if criminal_counts:
                    for count in criminal_counts:
                        rows.append(
                            {
                                **base,
                                "charge": count.get("name", ""),
                                "disposition": count.get("disposition", ""),
                                "status": count.get("status", ""),
                            }
                        )
                else:
                    rows.append({**base, "charge": "", "disposition": "", "status": ""})

        output_dir = Path(output_path)
        output_dir.mkdir(parents=True, exist_ok=True)

        stem = filename_stem or f"parties_{docket_id}"

        if output_format == "json":
            filepath = save_json(
                {"count": len(rows), "results": rows}, output_dir, filename_stem=stem
            )
        elif output_format == "csv":
            filepath = save_csv(rows, output_dir, filename_stem=stem)
        else:
            filepath = save_xlsx(rows, output_dir, filename_stem=stem)

        click.echo(f"✓ {len(parties)} parties, {len(rows)} charge rows")
        click.echo(f"✓ Saved to {filepath}")

    except Exception as e:
        click.echo(f"Error: {e}", err=True)
        raise SystemExit(1)
