#!/usr/bin/env python3
"""Profile a courtlistener-cli output file (JSON/CSV/XLSX) and print stats.

Usage: python result_stats.py <results-file> [--top N]

Prints: row count, column fill rates, value distributions for categorical
columns (court, status, type...), date ranges and per-year counts for date
columns, numeric min/max/mean. Stdlib + openpyxl only — no pandas.
"""

import csv
import json
import re
import statistics
import sys
from collections import Counter
from pathlib import Path

DATE_RE = re.compile(r"^\d{4}-\d{2}-\d{2}")


def load(path: Path) -> list[dict]:
    if path.suffix == ".json":
        data = json.loads(path.read_text())
        rows = data.get("results", data) if isinstance(data, dict) else data
        return [flatten(r) if isinstance(r, dict) else {"value": r} for r in rows]
    if path.suffix == ".csv":
        with open(path, newline="", encoding="utf-8") as f:
            return list(csv.DictReader(f))
    if path.suffix == ".xlsx":
        from openpyxl import load_workbook

        ws = load_workbook(path, read_only=True).active
        it = ws.iter_rows(values_only=True)
        header = next(it)
        return [dict(zip(header, row)) for row in it]
    sys.exit(f"Unsupported file type: {path.suffix}")


def flatten(d: dict, prefix: str = "") -> dict:
    out = {}
    for k, v in d.items():
        key = f"{prefix}.{k}" if prefix else k
        if isinstance(v, dict):
            out.update(flatten(v, key))
        elif isinstance(v, list):
            out[key] = json.dumps(v)
        else:
            out[key] = v
    return out


def profile(rows: list[dict], top: int = 10) -> None:
    n = len(rows)
    print(f"rows: {n}")
    if not n:
        return
    cols = list(rows[0].keys())
    print(f"columns: {len(cols)}\n")

    for col in cols:
        values = [r.get(col) for r in rows]
        filled = [v for v in values if v not in (None, "", "null")]
        if not filled:
            continue
        strs = [str(v) for v in filled]
        uniq = set(strs)
        line = f"{col}: {len(filled)}/{n} filled, {len(uniq)} unique"

        if all(DATE_RE.match(s) for s in strs[:50]):
            years = Counter(s[:4] for s in strs)
            line += f" | range {min(strs)[:10]} .. {max(strs)[:10]}"
            line += " | by year: " + ", ".join(
                f"{y}={c}" for y, c in sorted(years.items())
            )
        elif all(isinstance(v, (int, float)) for v in filled):
            nums = [float(v) for v in filled]
            line += f" | min={min(nums):g} max={max(nums):g} mean={statistics.fmean(nums):g}"
        elif len(uniq) <= max(top, n * 0.5) and len(uniq) < n:
            common = Counter(strs).most_common(top)
            line += " | top: " + ", ".join(f"{v[:40]}={c}" for v, c in common)
        print(line)


if __name__ == "__main__":
    args = sys.argv[1:]
    top = 10
    if "--top" in args:
        i = args.index("--top")
        top = int(args[i + 1])
        del args[i : i + 2]
    if not args:
        sys.exit(__doc__)
    profile(load(Path(args[0])), top)
